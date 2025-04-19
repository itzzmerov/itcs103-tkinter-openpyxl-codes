# main_app.py
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

# ========== Validation Function ==========
def validate_inputs():
    name = name_entry.get()
    email = email_entry.get()
    dob = dob_entry.get()

    if not name or not email or not dob:
        messagebox.showerror("Input Error", "All fields are required!")
        return False

    try:
        datetime.strptime(dob, "%Y-%m-%d")
    except ValueError:
        messagebox.showerror("Input Error", "Date must be in YYYY-MM-DD format!")
        return False

    return True

# ========== Save to Excel ==========
def save_to_excel():
    if not validate_inputs():
        return

    name = name_entry.get()
    email = email_entry.get()
    dob = dob_entry.get()

    wb = load_workbook("userdata.xlsx")
    ws = wb["UserData"]
    ws.append([name, email, dob])
    wb.save("userdata.xlsx")

    format_excel()  # Optional formatting
    messagebox.showinfo("Success", "Data saved successfully!")

    name_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    dob_entry.delete(0, tk.END)

# ========== Format Excel Sheet ==========
def format_excel():
    wb = load_workbook("userdata.xlsx")
    ws = wb["UserData"]

    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Format DOB column (col 3)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "YYYY-MM-DD"

    wb.save("userdata.xlsx")

# ========== View Excel Data ==========
def show_data():
    wb = load_workbook("userdata.xlsx")
    ws = wb["UserData"]

    data_window = tk.Toplevel(window)
    data_window.title("Stored User Data")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, value in enumerate(row):
            label = tk.Label(data_window, text=value, borderwidth=1, relief="solid", padx=6, pady=3)
            label.grid(row=i, column=j)

# ========== Tkinter UI ==========
window = tk.Tk()
window.title("User Data Form")

# Labels
tk.Label(window, text="Name").grid(row=0, column=0, padx=10, pady=5, sticky="w")
tk.Label(window, text="Email").grid(row=1, column=0, padx=10, pady=5, sticky="w")
tk.Label(window, text="Date of Birth (YYYY-MM-DD)").grid(row=2, column=0, padx=10, pady=5, sticky="w")

# Entry Fields
name_entry = tk.Entry(window, width=30)
email_entry = tk.Entry(window, width=30)
dob_entry = tk.Entry(window, width=30)

name_entry.grid(row=0, column=1, pady=5)
email_entry.grid(row=1, column=1, pady=5)
dob_entry.grid(row=2, column=1, pady=5)

# Buttons
tk.Button(window, text="Submit", command=save_to_excel, width=20, bg="#61AD4E", fg="white").grid(row=3, column=0, columnspan=2, pady=10)
tk.Button(window, text="View Stored Data", command=show_data, width=20, bg="#226710", fg="white").grid(row=4, column=0, columnspan=2)

window.mainloop()
