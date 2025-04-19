from openpyxl import load_workbook  

# Load an existing workbook
workbook = load_workbook("sample.xlsx")
sheet = workbook.active  # Get active sheet

# Read data from cells
print("Reading Excel File:")
for row in sheet.iter_rows(values_only=True):
    print(row)